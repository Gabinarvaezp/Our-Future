pip install pandas openpyxl matplotlib seaborn
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.drawing.image import Image
from openpyxl.chart import PieChart, Reference, BarChart, LineChart
from openpyxl.worksheet.dimensions import ColumnDimension
from datetime import datetime

class FinancialPlanner:
    def __init__(self):
        self.exchange_rate = 0.00025  # 1 COP = 0.00025 USD
        self.workbook = Workbook()
        self.setup_colors()
        self.months = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
                      'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        
    def setup_colors(self):
        self.colors = {
            'header_blue': '4472C4',
            'header_yellow': 'FFD700',
            'pastel_blue': 'B4D8E7',
            'pastel_yellow': 'FFF4BD',
            'pastel_green': 'C8E6C9',
            'pastel_pink': 'FFD1DC'
        }

    def create_annual_planner(self):
        # Crear hoja de resumen anual
        self.create_annual_summary()
        
        # Crear hojas mensuales
        for month in self.months:
            self.create_monthly_sheet(month)
        
        # Crear hojas adicionales
        self.create_savings_tracker()
        self.create_debt_control()
        self.create_travel_planner()

    def create_annual_summary(self):
        ws = self.workbook.active
        ws.title = "Resumen Anual"
        
        # Diseño del encabezado
        self.create_fancy_header(ws, "RESUMEN FINANCIERO ANUAL 2024", 1)
        
        # Columnas para el resumen
        headers = ['Mes', 'Ingresos Totales (USD)', 'Gastos Totales (USD)', 
                  'Ahorros (USD)', '% Gastos', '% Ahorro']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self.style_header_cell(cell)
            ws.column_dimensions[chr(64 + col)].width = 20

    def create_monthly_sheet(self, month):
        ws = self.workbook.create_sheet(month)
        
        # Configurar ancho de columnas
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20

        # Primera Quincena
        self.create_section_header(ws, f"{month} - Primera Quincena", 1, self.colors['header_blue'])
        
        # Ingresos Primera Quincena
        income_headers = ['Concepto', 'Monto', 'Moneda', 'USD Equivalente']
        row = 3
        for col, header in enumerate(income_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self.style_header_cell(cell)
        
        # Gastos Fijos Jorge
        row = 6
        self.create_section_header(ws, "Gastos Fijos Jorge (USD)", row, self.colors['pastel_yellow'])
        fixed_expenses_jorge = {
            'SchoolFirst': 400,
            'Cozy House': 400,
            'Cooper': 730,
            'Car Insurance': 140
        }
        
        # Gastos Fijos Gabby
        row = 12
        self.create_section_header(ws, "Gastos Fijos Gabby", row, self.colors['pastel_pink'])
        fixed_expenses_gabby = {
            'Casa': 400000,
            'Tarjeta Crédito': 500000,
            'Cozy House': 400000
        }

        # Segunda Quincena
        row = 18
        self.create_section_header(ws, f"{month} - Segunda Quincena", row, self.colors['header_yellow'])
        
        # Gastos Variables
        row = 24
        self.create_section_header(ws, "Gastos Variables", row, self.colors['pastel_green'])
        variable_expenses = {
            'Comida': '',
            'Transporte': '',
            'Entretenimiento': '',
            'Otros': ''
        }
        
        # Resumen Mensual
        row = 30
        self.create_section_header(ws, "Resumen Mensual", row, self.colors['header_blue'])
        
        # Gráficos
        self.create_monthly_charts(ws, row + 15)

    def create_savings_tracker(self):
        ws = self.workbook.create_sheet("Plan de Ahorro")
        
        self.create_fancy_header(ws, "PLANIFICADOR DE AHORRO", 1)
        
        # Meta Cartagena
        savings_goals = {
            'Hotel All Inclusive': 650,
            'Vuelos': 200,
            'Transporte': 80,
            'Cena Romántica': 100,
            'Actividades': 70
        }
        
        row = 3
        self.create_section_header(ws, "META: VIAJE A CARTAGENA", row, self.colors['pastel_blue'])
        total_goal = sum(savings_goals.values())
        
        # Tabla de metas
        headers = ['Concepto', 'Monto (USD)', 'Progreso', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row + 1, column=col, value=header)
            self.style_header_cell(cell)

    def create_debt_control(self):
        ws = self.workbook.create_sheet("Control de Deudas")
        
        self.create_fancy_header(ws, "CONTROL DE DEUDAS", 1)
        
        headers = ['Deuda', 'Monto Original', 'Interés', 'Pago Mensual', 'Saldo', 'Fecha Último Pago']
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self.style_header_cell(cell)

    def create_travel_planner(self):
        ws = self.workbook.create_sheet("Viaje Cartagena")
        
        self.create_fancy_header(ws, "PLANIFICADOR VIAJE CARTAGENA", 1)
        
        # Detalles del viaje
        trip_details = {
            'Fechas': '25-28 de [Mes]',
            'Hotel All Inclusive': 650,
            'Vuelos': 200,
            'Transporte Local': 80,
            'Cena Romántica': 100,
            'Actividades Planeadas': 'Atardecer, Playas, Tour Ciudad'
        }
        
        row = 3
        self.create_section_header(ws, "DETALLES DEL VIAJE", row, self.colors['pastel_blue'])

    def create_monthly_charts(self, ws, row):
        # Gráfico de Gastos
        pie = PieChart()
        pie.title = "Distribución de Gastos"
        data = Reference(ws, min_col=2, min_row=6, max_row=16)
        labels = Reference(ws, min_col=1, min_row=6, max_row=16)
        pie.add_data(data)
        pie.set_categories(labels)
        ws.add_chart(pie, f"F{row}")
        
        # Gráfico de Progreso de Ahorro
        line = LineChart()
        line.title = "Progreso de Ahorro"
        data = Reference(ws, min_col=2, min_row=30, max_row=35)
        labels = Reference(ws, min_col=1, min_row=30, max_row=35)
        line.add_data(data)
        line.set_categories(labels)
        ws.add_chart(line, f"F{row + 15}")

    def create_fancy_header(self, ws, title, row):
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(size=16, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=self.colors['header_blue'], 
                              end_color=self.colors['header_blue'], 
                              fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        
        border = Border(
            left=Side(style='medium', color=self.colors['header_yellow']),
            right=Side(style='medium', color=self.colors['header_yellow']),
            top=Side(style='medium', color=self.colors['header_yellow']),
            bottom=Side(style='medium', color=self.colors['header_yellow'])
        )
        cell.border = border

    def style_header_cell(self, cell):
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=self.colors['pastel_blue'], 
                              end_color=self.colors['pastel_blue'], 
                              fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    def save(self, filename=None):
        if filename is None:
            filename = f"Planificador_Financiero_{datetime.now().strftime('%Y%m%d')}.xlsx"
        self.workbook.save(filename)

# Crear y guardar el archivo
planner = FinancialPlanner()
planner.create_annual_planner()
planner.save()

print("¡Archivo Excel creado exitosamente!")
